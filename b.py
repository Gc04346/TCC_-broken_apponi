# encoding: utf-8
import decimal
import logging
import time
from typing import List

import traceback
from auditlog.registry import auditlog
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.core.validators import MaxValueValidator, MinValueValidator
from notifications.signals import notify
from post_office import mail

from music_system.settings.base import GOOGLE_STORAGE_DATASET_ID, GOOGLE_STORAGE_BACKUP_BUCKET
from music_system.settings.local import GOOGLE_STORAGE_PARENT
from ..settings import INVOICE_SETTINGS

import datetime
from django.db.models import F, QuerySet
from django.db import models
from django.db.models import Q, Sum
from django.db.models.signals import pre_delete, post_save, pre_save
from django.dispatch import receiver
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.models import User
# Create your models here.

from music_system.apps.contrib.models.base_model import BaseModel
from music_system.apps.contrib.validators import validate_file_max_50000, validate_only_positive_values
from music_system.apps.label_catalog.models import DSP, Provider, YoutubeAsset, Asset, Product, Holder, Artist, \
    ProductLegacyUPC, AssetLegacyISRC
from music_system.apps.label_reports.models.file_reader import FileReader

from ...contrib.api_helpers import default_query_assets_by_args
from ...contrib.log_helper import log_error, log_tests
from ...contrib.models.backup_model import BackupHandler
from music_system.apps.notifications_helper.notification_helpers import notify_on_telegram
from ...contrib.models.object_filterer import ObjectFilterer

CURRENCY_CHOICES = (
    ('BRL', _('Brazilian Real')), ('USD', _('United States Dollar')), ('EUR', _('Euro')),
    ('GBP', _('Great Britain Pound')),
)

CURRENCY_SYMBOLS = {
    'BRL': 'R$', 'USD': '$', 'EUR': '€', 'GBP': '£'
}

BATCH_STATUS_CHOICES = (
    ('OP', _('Open')),
    ('PR', _('Processing Items')),
    ('PRS', _('Items Processed')),
    ('PI', _('Processing Invoices')),
    ('PE', _('Processing ERROR')),
    ('CL', _('Closed')),
    ('PA', _('Payed')),
    ('BA', _('Backing up')),
    ('BS', _('Backup Success')),
    ('BE', _('Backup Error')),
)

REPORT_STATUS_CHOICES = (
    ('PEN', _('Pending')),
    ('VAL', _('Valid')),
    ('PRO', _('Processing')),
    ('PPE', _('Pending ID Asset')),
    ('SUC', _('Success')),
    ('ERR', _('Error')),
)

REPORT_TYPE_CHOICES = (
    ('DEEZ01', _('Deezer V1')),
    ('FCBK01', _('Facebook V1')),
    ('FCBK02', _('Facebook PMV V1')),
    ('FUGA01', _('Fuga V1')),
    ('GENERIC', _('Generic')),
    ('PAND01', _('Pandora V1')),
    ('SNDCLD01', _('SoundCloud V1')),
    ('SPTF01', _('Spotify V1')),
    ('TKTK01', _('TikTok V1')),
    ('RESS01', _('Resso V1')),
    ('YTAS01', _('Youtube Assets - Asset V1.0')),
    ('YTAS02', _('Youtube Assets - Asset V1.1')),
)

REPORT_ITEM_FORMAT_CHOICES = (
    ('AU', _('Audio')),
    ('VI', _('Video')),
)

INVOICE_STATUS_CHOICES = (
    ('OP', _('Open')), ('OH', _('On Hold')), ('CL', _('Closed/Payed')),
)

GENERIC_REPORT_FIELDS = (
    ('PRODUCT', _('REQUIRED')),
    ('DSP', _(
        'REQUIRED. Overrides the report DSP. The name must match the system database. Otherwise will be accounted as '
        'NULL.')),
    ('DSP_TRACK_ID', _('OPTIONAL.')),
    ('YOUTUBE_ASSET_ID', _('REQUIRED at least one of: YOUTUBE_ASSET_ID or ISRC or UPC')),
    ('ISRC', _('REQUIRED at least one of: YOUTUBE_ASSET_ID or ISRC or UPC')),
    ('UPC', _('REQUIRED at least one of: YOUTUBE_ASSET_ID or ISRC or UPC')),
    ('TITLE', _('OPTIONAL')),
    ('REFERENCE', _('OPTIONAL')),
    ('COUNTRY', _('OPTIONAL')),
    ('QUANTITY', _('REQUIRED')),
    ('VALUE', _('REQUIRED')),
    ('REPORT_START_DATE', _('OPTIONAL. Must be formated as a date. Otherwise will be accounted as NULL.')),
    ('REPORT_END_DATE', _('OPTIONAL. Must be formated as a date. Otherwise will be accounted as NULL.')),
)

INVOICE_ORDER_COLUMN_CHOICES = ['number', 'batch_name', 'holder_name', 'value_converted', 'tax_retention', 'commission',
                                'get_status_display', 'payment_date']


class Batch(BaseModel):
    """Batch is a class used to group reports for a specif date """

    class BackupStatusOptions(models.TextChoices):
        NOT_STARTED = ('NSTRT', _('Not Started'))
        IN_PROGRESS = ('INPRG', _('In Progress'))
        ERROR = ('ERROR', _('Error'))
        PENDING_VALIDATION = ('PDVAL', _('Pending Validation'))
        DONE = ('BDONE', _('Done'))
        ITEMS_DELETED = ('ITDEL', _('Items Deleted'))

    base_currency = models.CharField(verbose_name=_('Base Currency'), max_length=3, choices=CURRENCY_CHOICES,
                                     default='BRL')
    name = models.CharField(verbose_name=_('Name'), max_length=100, unique=True)
    status = models.CharField(verbose_name=_('Status'), max_length=3, choices=BATCH_STATUS_CHOICES, default='OP')
    errors = models.TextField(verbose_name=_('Errors'), null=True, blank=True)
    start_date = models.DateField(verbose_name=_('Start Date'))
    end_date = models.DateField(verbose_name=_('End Date'))
    invoice_due_date = models.DateField(verbose_name=_('Invoice Due Date'), null=True, blank=True)
    notes = models.TextField(verbose_name=_('Notes'), blank=True)
    backup_status = models.CharField(verbose_name=_('Backup Status'), max_length=5, choices=BackupStatusOptions.choices,
                                     default=BackupStatusOptions.NOT_STARTED)

    class Meta:
        verbose_name = _('Batch')
        verbose_name_plural = _('Batches')
        permissions = [
            ("distribute", "Can distribute the batch"),
            ("generate_invoices", "Can make the bathes invoices"),
        ]

    def __str__(self):
        """str method"""
        return f"{self.name} ({self.base_currency})"

    @classmethod
    def filter_objects(cls, searched_value: str, request_user: User, values_list_fields: list = None,
                       custom_query: Q = Q(), initial_queryset: QuerySet('Batch') = None) -> QuerySet:
        """Filtra os objetos da classe com base na string passada como parâmetro. Vide docstring do método chamado."""
        queryset = initial_queryset or cls.objects.all()
        search_fields = ['name', ]
        # Caso o usuario tenha passado uma query como parametro, o filtro sera feito com base nela apenas
        if custom_query:
            queryset = queryset.filter(custom_query)
        return ObjectFilterer.filter_objects(cls, searched_value, request_user.user_user_profile, search_fields,
                                             queryset, values_list_fields)

    @property
    def raw_commission_profit(self):
        return Invoice.objects.filter(batch_id=self.id).aggregate(Sum('value_commission_converted')).get(
            'value_commission_converted__sum', '-')

    raw_commission_profit.fget.short_description = _('Raw Profit (Comm.)')

    @property
    def raw_discounted_profit(self):
        return sum(invoice.net_distributed_value for invoice in Invoice.objects.filter(batch_id=self.id))

    raw_discounted_profit.fget.short_description = _('Raw Profit (Disc.)')

    def make_invoices(self):
        """Translates the temp items to report items, converting the value.
        """
        # make items
        if self.status not in ['PRS', 'CL']:
            return
        self.status = 'PI'
        self.errors = ''
        self.save()
        try:
            holders = Holder.objects.all()
            for holder in holders:
                # faz invoice
                report_items = ReportItem.objects.filter(holder=holder, report__batch=self)
                if len(report_items) > 0:
                    invoice_value_converted = report_items.aggregate(total=Sum('value_holder_converted'))[
                        'total']
                    value_distributor_converted = report_items.aggregate(total=Sum('value_converted_distributor'))[
                        'total']
                    tax_retention = invoice_value_converted * holder.tax_retention / 100
                    # todo fix this
                    try:
                        invoice = self.invoice_set.get(holder=holder)
                    except ObjectDoesNotExist:
                        import uuid
                        number = uuid.uuid4()
                        invoice = Invoice()
                        invoice.number = number

                    invoice.holder = holder
                    invoice.batch = self
                    invoice.value_distributor_converted = round(value_distributor_converted, 2)
                    invoice_value_without_advances = round(invoice_value_converted - tax_retention, 2)
                    invoice.value_converted = invoice_value_without_advances
                    invoice.status = 'OP'
                    invoice.tax_retention = round(tax_retention, 2)

                    invoice.bank_data = holder.bank_data
                    invoice.legal_document = holder.legal_document
                    invoice.legal_name = holder.legal_name

                    invoice.save()

                    # Depois que eu sei o valor da invoice eu posso criar as amortizações, porque eu preciso saber o
                    # valor total da invoice antes de fazer isso.
                    holder_advances = Advance.objects.filter(holder=holder,
                                                             will_be_discounted=True).order_by('advance_date')
                    # Confere se o titular tem adiantamentos
                    advances_qtd = len(holder_advances)
                    i = 0  # Contador que vai ser usado pra iterar sobre os adiantamentos
                    total_amortization_value = 0  # Valor que será descontado da ND referente a adiantamento(s)
                    while bool((i < advances_qtd) and (total_amortization_value < invoice_value_without_advances)):
                        advance = holder_advances[i]
                        i += 1
                        # O valor abaixo representa o quanto ainda pode ser amortizado dessa invoice
                        max_amortization_value = invoice.value_converted - total_amortization_value
                        # Cria uma amortização, respeitando o valor total a ser recebido pelo titular nessa ND
                        installment = advance.create_advance_invoice_relation(invoice, max_amortization_value)
                        # Se não tiver sido criado nenhuma amortização, pula pra próx iteração, pq pode não ter sido
                        # criada por qualquer um dos fatores listados em Advance.has_to_be_discounted(), e isso tem
                        # que ser respeitado no algoritmo.
                        if not installment:
                            continue
                        total_amortization_value += installment.amount

                    from ..tasks import make_invoice_files
                    make_invoice_files.apply_async((invoice.id,), countdown=1)
            self.status = 'CL'
            self.save()
        except Exception as e:
            import traceback
            traceback.print_stack()
            log_error(e)
            log_tests(e)
            self.status = 'PE'
            self.errors = _('Unknown error.')
            self.save()

    def make_invoices_files(self):
        """Translates the temp items to report items, converting the value.
        """
        for invoice in self.invoice_set.all():
            # invoice.make_files()
            from ..tasks import make_invoice_files
            make_invoice_files.apply_async((invoice.id,), countdown=1)

    def distribute_batch(self):
        """Translates the temp items to report items, converting the value.
        """
        # make items
        if self.status == 'OP':
            self.status = 'PR'
            self.errors = ''
            self.save()
            reports = self.report_set.all()
            try:
                for report in reports:
                    gross_revenue_distributed = 0
                    net_revenue_distributed = 0
                    net_income_distributed = 0
                    items_distributed = 0

                    report.status_distribution = 'PRO'
                    report.save()
                    for item in report.reporttempitemasset_set.all():
                        values_dist = item.distribute_item(report)
                        gross_revenue_distributed += values_dist['gross_revenue_distributed']
                        net_revenue_distributed += values_dist['net_revenue_distributed']
                        net_income_distributed += values_dist['net_income_distributed']
                        items_distributed += values_dist['items_distributed']
                        item.delete()
                    for item in report.reporttempitemyoutube_set.all():
                        values_dist = item.distribute_item(report)
                        gross_revenue_distributed += values_dist['gross_revenue_distributed']
                        net_revenue_distributed += values_dist['net_revenue_distributed']
                        net_income_distributed += values_dist['net_income_distributed']
                        items_distributed += values_dist['items_distributed']
                        item.delete()
                    report.gross_revenue_distributed = gross_revenue_distributed
                    report.net_revenue_distributed = net_revenue_distributed
                    report.net_income_distributed = net_income_distributed
                    report.items_distributed = items_distributed
                    report.status_distribution = 'SUC'
                    report.save()

                self.status = 'PRS'
                self.save()
            except Exception as e:
                log_error(e)
                self.status = 'PE'
                self.errors = _('Unknown error.')
                self.save()
            str1 = _('finished distributing with status')
            notify_on_telegram('financeiro',
                               f"{_('Batch')} {self} {str1}: {self.get_status_display()}.")

    @property
    def can_change_batch(self):
        return self.can_change_batch_any(self)

    @property
    def can_distribute_batch(self):
        """
            O lote pode ser distribuído se não houverem relatórios com status diferente de sucesso e seu status estiver
             como Aberto.
        """
        return self.report_set.filter(~Q(status='SUC')).count() == 0 and self.status in ['OP']

    @property
    def can_generate_invoices(self):
        return self.status in ['PRS', 'CL']

    @staticmethod
    def can_change_batch_any(obj):
        """Checks if the object can be changed or deleted. Closed batches cannot be changed but they can be deleted.
        """
        # noinspection PyBroadException
        try:
            return obj.status not in [Batch.get_payed_status_code()] + Batch.get_backup_status_codes()
        except Exception:
            return False

    @staticmethod
    def get_open_status_code():
        return 'OP'

    @staticmethod
    def get_closed_status_code():
        return 'CL'

    @staticmethod
    def get_payed_status_code():
        return 'PA'

    @staticmethod
    def get_backup_status_codes():
        return ['BA', 'BE', 'BS']

    @staticmethod
    def get_statuses_available_to_holder_list():
        return [Batch.get_closed_status_code(), Batch.get_payed_status_code()] + Batch.get_backup_status_codes()

    @property
    def gross_revenue_converted(self):
        # noinspection PyBroadException
        try:
            return round(self.report_set.all().aggregate(
                total=Sum(F('gross_revenue_distributed') * F('currency_price')))['total'], 2)
        except Exception:
            return 0.00

    def get_gross_revenue_converted(self):
        return self.gross_revenue_converted

    get_gross_revenue_converted.short_description = '{} - {}'.format(_('Gross Rev.'), _('Base Currency'))

    @property
    def net_revenue_converted(self):
        # noinspection PyBroadException
        try:
            return round(self.report_set.all().aggregate(
                total=Sum(F('net_revenue_distributed') * F('currency_price')))['total'], 2)
        except Exception:
            return 0.00

    def get_net_revenue_converted(self):
        return self.net_revenue_converted

    get_net_revenue_converted.short_description = '{} - {}'.format(_('Net Rev.'), _('Base Currency'))

    @property
    def net_income_converted(self):
        # noinspection PyBroadException
        try:
            return round(self.report_set.all().aggregate(
                total=Sum(F('net_income_distributed') * F('currency_price')))['total'], 2)
        except Exception:
            return 0.00

    def get_net_income_converted(self):
        return self.net_income_converted

    get_net_income_converted.short_description = '{} - {}'.format(_('Net Income'), _('Base Currency'))

    def get_can_distribute(self):
        """Split and concatenates all errors messages
        """
        if not self.can_distribute_batch or self.id is None:
            return _('N/A') if self.status != 'OP' else format_html('<span class="icon-cross"></span>')

        button = f"""<a href="{reverse('label_reports:reports.confirm_distribution', args=[self.id])}" class="default">{_('Distribute')}</a>"""

        return format_html('<span class="icon-tick"></span>' + button)

    get_can_distribute.short_description = _('Distribute')

    def get_can_make_invoices(self):
        """Split and concatenates all errors messages
        """
        if self.status not in ['PRS', 'CL']:
            return _('N/A') if self.status != 'PA' else format_html('<span class="icon-cross"></span>')

        button = f"""<a href="{reverse('label_reports:reports.make_invoices', args=[self.id])}" class="default">{_('Make Invoices')}</a>"""

        return format_html('<span class="icon-tick"></span>' + button)

    get_can_make_invoices.short_description = _('Make Invoices')

    def get_can_make_invoices_files(self):
        """Split and concatenates all errors messages
        """
        if self.status not in ['PRS', 'CL']:
            return _('N/A') if self.status != 'PA' else format_html('<span class="icon-cross"></span>')

        button = f"""<a href="{reverse('label_reports:reports.make_invoices_files', args=[self.id])}" class="default">{_('Make Invoices Files')}</a>"""

        return format_html('<span class="icon-tick"></span>' + button)

    get_can_make_invoices_files.short_description = _('Make Invoices Files')

    def get_invoices_summary(self):
        if self.status not in ['PA', 'CL']:
            return _('N/A')
        return format_html(
            f"""<span class="icon-tick"></span><a href="{reverse('label_reports:reports.get_invoices_summary', args=[self.id])}" class="default">{_('Download')}</a> """)

    get_invoices_summary.short_description = _('Get Invoices Summary')

    def get_invoices_pdf(self):
        if self.status not in ['PA', 'CL']:
            return _('N/A')
        button = f"""<a href="{reverse('label_reports:reports.get_invoices_pdf', args=[self.id])}" class="default">{_('Download')}</a>"""

        return format_html('<span class="icon-tick"></span>' + button)

    get_invoices_pdf.short_description = _('Get Invoices PDFs')

    def get_invoices_all_files(self):
        if self.status not in ['PA', 'CL']:
            return _('N/A')
        button = f"""<a href="{reverse('label_reports:reports.get_invoices_all_files', args=[self.id])}" class="default">{_('Download')}</a>"""

        return format_html('<span class="icon-tick"></span>' + button)

    get_invoices_all_files.short_description = _('Get All Invoices Files')

    def make_backup_url(self):
        """ Url para chamada de backup """
        # Só pode fazer backup se todas as traduções estiverem preenchidas
        if DSPProductCodeTranslation.objects.only('translation').filter(
                Q(translation='') | Q(translation__isnull=True)).exists():
            return _('It was not possible to perform the backup because there are incomplete translations in the db.')
        # Só pode backupar se o lote estiver válido
        if self.status not in ['PA', 'CL', 'BE'] or self.backup_status not in [self.BackupStatusOptions.NOT_STARTED,
                                                                               self.BackupStatusOptions.ERROR]:
            return _('N/A')
        return format_html(
            '<span class="icon-tick"></span><a ' +
            f'href="{reverse("label_reports:reports.backup_batch", args=[self.id])}" class="default">{_("Backup")}</a>')

    def delete_items_url(self):
        """ Url para chamada de apagar report itens do lote """
        # Só pode apagar os itens se o status do lote estiver como "feito"
        if self.backup_status != self.BackupStatusOptions.DONE:
            return '-'
        return format_html(
            '<span class="icon-tick"></span><a ' +
            f'href="{reverse("label_reports:reports.delete_report_items", args=[self.id])}" class="default">{_("Delete")}</a>')

    @staticmethod
    def bigquery_dataframe_data_types():
        """DataTypes para backup no BigQuery"""
        return {
            'titular': 'string',
            'titular_id': 'int',
            'lote': 'string',
            'nome_do_lote': 'string',
            'lote_id': 'int',
            'artista': 'string',
            'artista_feat': 'string',
            'share_titular': 'float',
            'upc': 'string',
            'isrc': 'string',
            'youtube_asset_id': 'string',
            'titulo_produto': 'string',
            'titulo': 'string',
            'formato': 'string',
            'dsp': 'string',
            'tipo_de_uso': 'string',
            'quantidade': 'int',
            'pais': 'string',
            'valor_total': 'float',
            'valor_total_real': 'float',
            'share_titular_repasse': 'float',
            'valor_titular': 'float',
            'valor_titular_real': 'float',
            'share_distribuidor': 'float',
            'valor_distribuidor': 'float',
            'valor_distribuidor_real': 'float',
            'data_inicial': 'datetime64[ns]',
            'data_final': 'datetime64[ns]',
        }

    @staticmethod
    def bigquery_table_schema():
        """Table Schema para backup no BigQuery"""
        return [
            {'name': 'titular', 'type': 'STRING'},
            {'name': 'titular_id', 'type': 'INTEGER'},
            {'name': 'lote', 'type': 'STRING'},
            {'name': 'nome_do_lote', 'type': 'STRING'},
            {'name': 'lote_id', 'type': 'INTEGER'},
            {'name': 'artista', 'type': 'STRING'},
            {'name': 'artista_feat', 'type': 'STRING'},
            {'name': 'share_titular', 'type': 'FLOAT'},
            {'name': 'upc', 'type': 'STRING'},
            {'name': 'isrc', 'type': 'STRING'},
            {'name': 'youtube_asset_id', 'type': 'STRING'},
            {'name': 'titulo_produto', 'type': 'STRING'},
            {'name': 'titulo', 'type': 'STRING'},
            {'name': 'formato', 'type': 'STRING'},
            {'name': 'dsp', 'type': 'STRING'},
            {'name': 'tipo_de_uso', 'type': 'STRING'},
            {'name': 'quantidade', 'type': 'INTEGER'},
            {'name': 'pais', 'type': 'STRING'},
            {'name': 'valor_total', 'type': 'FLOAT'},
            {'name': 'valor_total_real', 'type': 'FLOAT'},
            {'name': 'share_titular_repasse', 'type': 'FLOAT'},
            {'name': 'valor_titular', 'type': 'FLOAT'},
            {'name': 'valor_titular_real', 'type': 'FLOAT'},
            {'name': 'share_distribuidor', 'type': 'FLOAT'},
            {'name': 'valor_distribuidor', 'type': 'FLOAT'},
            {'name': 'valor_distribuidor_real', 'type': 'FLOAT'},
            {'name': 'data_inicial', 'type': 'DATETIME'},
            {'name': 'data_final', 'type': 'DATETIME'},
        ]

    def make_backup(self, iteration_number, total_iterations=None):
        """ Faz o backup e salva no google cloud storage """
        if not total_iterations or iteration_number == 1:
            # Só manda essa notificação na primeira iteração pra evitar spam
            notify_on_telegram('dev', bytes.decode(b'\xF0\x9F\x92\xB9',
                                                   'utf-8') + ' Backup de itens financeiros prestes a iniciar.')
        translation_codes = {f'{translation.dsp}-{translation.code}': translation.translation for translation in
                             DSPProductCodeTranslation.objects.all()}
        if any(invalid_value in translation_codes.values() for invalid_value in [None, '']):
            error_msg = _(
                'It was not possible to perform the backup because there are incomplete translations in the db.')
            log_error(error_msg)
            notify_on_telegram('dev', error_msg)
            return False
        # make_base_lists
        holders = {holder.id: holder.name for holder in Holder.objects.only('id', 'name').all()}
        report_ids = [report.id for report in Report.objects.only('id').filter(batch_id=self.id)]
        chunk_size = 200_000
        report_items_ids = ReportItem.objects.only('id').values_list('id', flat=True).filter(
            report_id__in=report_ids,
            backup_done=False).order_by('id')
        total_items = report_items_ids.count()
        report_items_ids = report_items_ids[:chunk_size]
        report_items = ReportItem.objects.filter(id__in=report_items_ids)
        headers = ['titular', 'titular_id', 'lote', 'nome_do_lote', 'lote_id', 'artista', 'artista_feat',
                   'share_titular',
                   'upc', 'isrc', 'youtube_asset_id', 'titulo_produto', 'titulo', 'formato', 'dsp', 'tipo_de_uso',
                   'quantidade', 'pais', 'valor_total', 'valor_total_real', 'share_titular_repasse', 'valor_titular',
                   'valor_titular_real', 'share_distribuidor', 'share_extra_distribuidor', 'valor_distribuidor',
                   'valor_distribuidor_real', 'data_inicial', 'data_final']
        data = [[holders[report_item.holder_id],
                 report_item.holder_id,
                 self.name[:7],
                 self.name,
                 self.id,
                 report_item.artists_names,
                 report_item.artists_feat_names,
                 report_item.share_holder,
                 report_item.upc,
                 report_item.isrc,
                 report_item.youtube_asset_asset_id,
                 report_item.product_title,
                 report_item.title,
                 report_item.format,
                 report_item.dsp,
                 translation_codes.get(f"{report_item.dsp}-{report_item.product_dsp}", 'N/F'),  # N/F - Not Found
                 report_item.quantity,
                 report_item.country,
                 report_item.value_holder + report_item.value_distributor,
                 report_item.value_holder_converted + report_item.value_converted_distributor,
                 report_item.holder_item_share,
                 report_item.value_holder,
                 report_item.value_holder_converted,
                 report_item.rate_distributor,
                 report_item.transfer_dist_share,
                 report_item.value_distributor,
                 report_item.value_converted_distributor,
                 report_item.start_date,
                 report_item.end_date]
                for report_item in report_items]
        import math
        iterations_needed = total_iterations or math.ceil(total_items / chunk_size)
        handler = BackupHandler()
        if success := handler.perform_backup(headers, data, GOOGLE_STORAGE_DATASET_ID, GOOGLE_STORAGE_BACKUP_BUCKET,
                                             f'report_item_{iteration_number}_of_{iterations_needed}',
                                             f'financial_batch/{self.name}_{iteration_number}_of_{iterations_needed}.csv',
                                             'itens_financeiro'):
            report_items.update(backup_done=True)
            if total_items > chunk_size:  # Tem mais itens do que o tamanho do bloco?
                self.backup_status = self.BackupStatusOptions.IN_PROGRESS
                self.save()
                from music_system.apps.label_reports.tasks import make_backup as backup_task
                backup_task.apply_async((self.id, iteration_number + 1, iterations_needed), countdown=1)
                return True  # Indica que será necessária outra iteração da task
            else:
                self.backup_status = self.BackupStatusOptions.PENDING_VALIDATION
                self.save()
                return False  # Indica que todos os registros couberam no bloco e foram backupados
        else:
            self.backup_status = self.BackupStatusOptions.ERROR
            self.save()
            return False

    def delete_report_items(self):
        """ Deleta todos os itens de relatório deste lote """
        ReportItem.objects.filter(backup_done=True,
                                  report_id__in=self.report_set.only('id').all().values_list('id', flat=True)).delete()


# noinspection PyUnusedLocal
def get_report_file_path(instance, filename, folder='uploads'):
    import os
    import uuid
    ext = filename.split('.')[-1]
    filename = f"{uuid.uuid4()}.{ext}"
    return os.path.join(folder, filename)


class Report(BaseModel):
    """Self-explained. """
    currency = models.CharField(verbose_name=_('Currency'), max_length=3, choices=CURRENCY_CHOICES,
                                default='BRL')
    currency_price = models.DecimalField(verbose_name=_('Currency Price'), default=1.00, max_digits=15,
                                         decimal_places=5)
    name = models.CharField(verbose_name=_('Name'), max_length=1000)
    batch = models.ForeignKey(verbose_name=_('Batch'), to=Batch, on_delete=models.CASCADE,
                              limit_choices_to={'status': Batch.get_open_status_code()})
    dsp = models.ForeignKey(verbose_name=_('DSP'), to=DSP, on_delete=models.CASCADE,
                            limit_choices_to={'active': True})
    provider = models.ForeignKey(verbose_name=_('Provider'), to=Provider, on_delete=models.CASCADE,
                                 limit_choices_to={'active': True})
    multiplier = models.DecimalField(verbose_name=_('Multiplier'),
                                     help_text=_('Value that is going to be multiplied by each item for adjustment'),
                                     default=1.00000, decimal_places=5, max_digits=7,
                                     validators=[validate_only_positive_values])
    gross_revenue_expected = models.DecimalField(verbose_name=_('Gross Rev. - Expected'), max_digits=15,
                                                 decimal_places=2)
    gross_revenue_distributed = models.DecimalField(verbose_name=_('Gross Rev. - Distributed'), max_digits=15,
                                                    decimal_places=2, default=0.00)
    gross_revenue = models.DecimalField(verbose_name=_('Gross Rev.'), default=0.00, max_digits=15, decimal_places=2)
    net_revenue_expected = models.DecimalField(verbose_name=_('Net Rev. - Expected'), max_digits=15,
                                               decimal_places=2)
    net_revenue = models.DecimalField(verbose_name=_('Net Rev.'), default=0.00, max_digits=15, decimal_places=2)
    net_revenue_distributed = models.DecimalField(verbose_name=_('Net Rev. - Distributed'), default=0.00,
                                                  max_digits=15,
                                                  decimal_places=2)
    provider_fees = models.DecimalField(verbose_name=_('Provider Fees'), default=0.00, max_digits=15,
                                        decimal_places=3)
    net_income_distributed = models.DecimalField(verbose_name=_('Net Income'), default=0.00, max_digits=15,
                                                 decimal_places=2)
    quantity = models.IntegerField(verbose_name=_('Quantity'), default=0)
    items = models.IntegerField(verbose_name=_('Items'), default=0)
    items_distributed = models.IntegerField(verbose_name=_('Items After Distribution'), default=0)
    status = models.CharField(verbose_name=_('Status'), max_length=10, choices=REPORT_STATUS_CHOICES, default='PEN')
    errors = models.TextField(verbose_name=_('Errors'), null=True, blank=True)
    status_distribution = models.CharField(verbose_name=_('Status Distribution'), max_length=10,
                                           choices=REPORT_STATUS_CHOICES,
                                           default='PEN')
    notes = models.TextField(verbose_name=_('Notes'), blank=True)
    start_date = models.DateField(verbose_name=_('Start Date'), null=True, blank=True)
    end_date = models.DateField(verbose_name=_('End Date'), null=True, blank=True)

    class Meta:
        verbose_name = _('Financial Report')
        verbose_name_plural = _('Financial Reports')

    def __str__(self):
        """str method"""
        return "{} ({})".format(self.name, self.currency_for_humans)

    @property
    def currency_for_humans(self):
        return "{} - {}".format(self.currency, round(self.currency_price, 2))

    @property
    def gross_revenue_nominal_converted(self):
        return round(self.gross_revenue_expected * self.currency_price, 2)

    @property
    def gross_revenue_distributed_converted(self):
        return round(self.gross_revenue_distributed * self.currency_price, 2)

    @property
    def gross_revenue_converted(self):
        return round(self.gross_revenue * self.currency_price, 2)

    @property
    def net_revenue_converted(self):
        return round(self.net_revenue * self.currency_price, 2)

    @property
    def provider_fees_converted(self):
        return round(self.provider_fees * self.currency_price, 2)

    @property
    def can_delete_report(self):
        """Checks if the object can be changed. Closed batches cannot be changed but they can be deleted.
        """
        return self.can_delete_report_any(self)

    @staticmethod
    def can_delete_report_any(obj):
        """Checks if the object can be changed. Closed batches cannot be changed but they can be deleted.
        """
        # noinspection PyBroadException
        try:
            return obj.status in ['SUC', 'ERR', 'PPE'] and obj.status_distribution in ['SUC', 'ERR', 'PEN']
        except Exception:
            return False

    @staticmethod
    def get_blank_generic_file():
        from music_system.apps.contrib.bulk_helper import get_blank_generic_file_from_fields
        return get_blank_generic_file_from_fields(GENERIC_REPORT_FIELDS)

    # todo make function to check if batch is payed when marking a report as payed
    def process_file(self):
        """ Processa os arquivos financeiros, produzindo itens temporários de DSP e Youtube. Durante o processamento, o
            status do relatório fica como "processando" (PRO). Ao final desse processo, o status do relatório pode ir
            para "sucesso" (SUC) ou "pendente identificação" (PPE), que indica que algum item do relatório não está pre-
            sente no nosso banco de dados (ou seja, não existe nenhum produto, asset ou youtubeasset correspondente).
        """
        if self.status == 'VAL':
            self.status = 'PRO'
            self.save()
            data = self.translate_file_to_dict()
            dsps_list = {}
            for dsp in DSP.objects.all():
                dsps_list[dsp.id] = dsp.name
            assets_list = {}
            for asset in Asset.objects.all():
                assets_list[asset.isrc] = asset.id
            for asset in AssetLegacyISRC.objects.all():
                assets_list[asset.isrc] = asset.asset_id
            products_list = {}
            for product in Product.objects.all():
                products_list[product.upc] = product.id
            for product in ProductLegacyUPC.objects.all():
                products_list[product.upc] = product.product_id
            youtube_assets_list = {}
            for youtube_asset in YoutubeAsset.objects.all():
                youtube_assets_list[youtube_asset.asset_id] = youtube_asset.id
            metrics_items = 0
            metrics_quantity = 0
            metrics_gross_revenue = 0
            metrics_net_revenue = 0
            metrics_provider_fees = 0
            temp_assets = []
            temp_youtube_assets = []
            count_pending_id = 0
            for linha, item in enumerate(data['items']):
                item_value = item['value']
                metrics_items += 1
                metrics_quantity += item['quantity']
                asset_kwargs = {
                    'dsp': str(dsps_list[item['dsp_id']])[:100],
                    'format': item['format'],
                    'report_id': self.id,
                    'upc': item['upc'],
                    'youtube_asset_asset_id': item['youtube_asset_id'],
                    'isrc': item['isrc'],
                    'value': item_value,
                    'quantity': item['quantity'],
                    'reference': item['reference'],
                    'country': item['country'],
                    'product_dsp': item['product_dsp'],
                    'start_date': item['start_date'],
                    'end_date': item['end_date']
                }
                if item['format'] == 'VI':
                    try:
                        asset_kwargs['youtube_asset_id'] = youtube_assets_list[item['youtube_asset_id']]
                    except KeyError:
                        count_pending_id += 1
                    # O ajuste da retenção de imposto só acontece em itens de Youtube com country US
                    # Esse ajuste não existe mais aparentemente
                    # if item.get('country') == 'US':
                    #     item_tax_retention = item_value * decimal.Decimal(0.3)
                    #     item_value = item_value - item_tax_retention
                    #     asset_kwargs['value'] = item_value
                    #     asset_kwargs['tax_retention'] = item_tax_retention
                    temp_youtube_assets.append(ReportTempItemYoutube(**asset_kwargs))
                else:
                    count_pending_id_product_and_asset = 0
                    try:
                        asset_kwargs['product_id'] = products_list[str(item['upc'])]
                    except KeyError:
                        count_pending_id_product_and_asset += 1
                    try:
                        asset_kwargs['asset_id'] = assets_list[item['isrc']]
                    except KeyError:
                        count_pending_id_product_and_asset += 1
                    if count_pending_id_product_and_asset == 2:
                        count_pending_id += 1
                    temp_assets.append(ReportTempItemAsset(**asset_kwargs))

                metrics_gross_revenue += item_value
                item_net_value = round(item_value * decimal.Decimal(1 - (self.provider.share / 100)), 5)
                metrics_net_revenue += item_net_value
                metrics_provider_fees += (item_value - item_net_value)

            if len(temp_youtube_assets) > 0:
                ReportTempItemYoutube.objects.bulk_create(temp_youtube_assets, batch_size=5000)
            if len(temp_assets) > 0:
                ReportTempItemAsset.objects.bulk_create(temp_assets, batch_size=5000)

            if count_pending_id > 0:
                self.status = 'PPE'
            else:
                self.status = 'SUC'
            self.quantity = metrics_quantity
            self.net_revenue = metrics_net_revenue
            self.gross_revenue = metrics_gross_revenue
            self.provider_fees = metrics_provider_fees
            self.items = metrics_items
            self.multiplier = data.get('multiplier', 1)
            self.save()
            # send_notification(author=self,
            #                   recipients=User.objects.filter(user_user_profile__is_staff_financial=True),
            #                   verb=_('was processed'),
            #                   url='javascript:void(0)')  # todo sera que vai prestar?

    def validate_file(self):
        """ Valida os dados lidos dos arquivos de relatório. Ao final, marca o relatório como VAL ou como ERR. Caso
            esteja válido, a tarefa de processar os arquivos financeiros é colocada na fila.
        """
        data = self.translate_file_to_dict()
        errors_count = 0
        errors_messages = []
        if data['status'] == 'error':
            errors_count += 1
            errors_messages.append(str(data['message']))
        if len(data['items']) <= 0 and data['status'] != 'error':
            errors_count += 1
            errors_messages.append(_('Empty report. Nothing was processed.'))
        elif self.status == 'PEN' or self.status == 'ERR':
            total = len(data["items"])
            for index, item in enumerate(data['items'], start=1):
                line_num_text = _('Line: {}.').format(index + 1)

                # check if item has at least 1 identifier
                if item['upc'] == item['isrc'] == item['youtube_asset_id'] == '':
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _('At least one of: YOUTUBE_ASSET_ID or ISRC or UPC')))
                if item['dsp_id'] is None:
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'DSP not found. If you used the Generic type, please use an existing DSP name, otherwise '
                            'contact support and send this message.')))

                if item['product_dsp'] is None or item['product_dsp'] == '':
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'Product (item type of use) is required.')))
                if item['format'] is None or item['format'] == '':
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'Format (audio or video) is required. Please contact support and send this message.')))
                if item['value'] is None:
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'Value is required but can be 0.')))
                if item['quantity'] is None:
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'Quantity is required but can be 0.')))
                if not isinstance(item['start_date'], datetime.date) and item['start_date'] is not None:
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'Start Date not valid. Please contact support and send this message.')))
                if not isinstance(item['end_date'], datetime.date) and item['end_date'] is not None:
                    errors_count += 1
                    errors_messages.append(
                        '{} {}'.format(line_num_text, _(
                            'End Date not valid. Please contact support and send this message.')))
            if errors_count == 0:
                self.status = 'VAL'
                from ..tasks import process_report_task
                process_report_task.apply_async((self.id,),
                                                eta=timezone.now() + timezone.timedelta(seconds=1))
        if errors_count > 0:
            self.status = 'ERR'
            self.errors = '|'.join(errors_messages)
        self.save()

    # noinspection DuplicatedCode
    @staticmethod
    def recheck_items_match(product_ids: list = "", asset_ids: list = "", youtube_asset_ids: list = ""):
        items_updated_reports = {}
        assets_list = {}
        for asset in Asset.objects.filter(id__in=asset_ids):
            assets_list[asset.isrc] = asset.id
        for asset in AssetLegacyISRC.objects.filter(asset_id__in=asset_ids):
            assets_list[asset.isrc] = asset.asset_id
        products_list = {}
        for product in Product.objects.filter(id__in=product_ids):
            products_list[str(product.upc)] = product.id
        for product in ProductLegacyUPC.objects.filter(product_id__in=product_ids):
            products_list[product.upc] = product.product_id
        youtube_assets_list = {}
        for youtube_asset in YoutubeAsset.objects.filter(id__in=youtube_asset_ids):
            youtube_assets_list[str(youtube_asset.asset_id)] = youtube_asset.id
        if len(youtube_assets_list) > 0 or len(products_list) > 0 or len(assets_list) > 0:
            pending_assets_list = ReportTempItemAsset.objects.filter(
                Q(Q(product_id__isnull=True) | Q(asset_id__isnull=True)))
            pending_youtube_assets_list = ReportTempItemYoutube.objects.filter(youtube_asset_id__isnull=True)

            for item in pending_assets_list:
                count_pending_id_product_and_asset = 0
                try:
                    item.product_id = products_list[str(item.upc)]
                except KeyError:
                    count_pending_id_product_and_asset += 1
                try:
                    item.asset_id = assets_list[item.isrc]
                except KeyError:
                    count_pending_id_product_and_asset += 1
                if count_pending_id_product_and_asset < 2:
                    item.save()
                    items_updated_reports[item.report_id] = True

            for item in pending_youtube_assets_list:
                if item.format == 'VI':
                    count_pending_id = 0
                    try:
                        item.youtube_asset_id = youtube_assets_list[item.youtube_asset_asset_id]
                    except KeyError:
                        count_pending_id += 1
                    if count_pending_id == 0:
                        item.save()
                        items_updated_reports[item.report_id] = True

        for report in Report.objects.filter(status='PPE'):
            try:

                missing_id = report.reporttempitemasset_set.filter(product_id__isnull=True,
                                                                   asset_id__isnull=True).count()
                missing_id += report.reporttempitemyoutube_set.filter(
                    youtube_asset_id__isnull=True).count()
                if missing_id == 0:
                    report.status = 'SUC'
                    report.save()
            except ObjectDoesNotExist:
                pass

    def translate_file_to_dict(self):
        file_reader = FileReader()
        return file_reader.translate_file_to_dict_any(self.reportfile_set.all(), (
            self.currency, self.dsp_id, self.gross_revenue_expected, self.start_date, self.end_date))

    def get_errors_messages_html(self):
        from music_system.apps.contrib.bulk_helper import get_errors_messages_html
        return get_errors_messages_html(self.errors)

    get_errors_messages_html.short_description = _('Errors')


class ReportFile(BaseModel):
    file = models.FileField(verbose_name=_('File'), help_text=_('Files are kept for 60 days. Max size is 50mb.'),
                            upload_to=get_report_file_path, validators=[validate_file_max_50000], max_length=200)
    report = models.ForeignKey(to=Report, on_delete=models.CASCADE, verbose_name=_('Financial Report'))
    report_type = models.CharField(verbose_name=_('Type'), max_length=25, choices=REPORT_TYPE_CHOICES)

    class Meta:
        verbose_name = _('Financial Report File')
        verbose_name_plural = _('Financial Report Files')

    def __str__(self):
        return f'{self.report}  - {self.file}'


class ReportTempItem(BaseModel):
    """Mother class to store report items before distribution.
    """

    dsp = models.CharField(verbose_name=_('DSP'), max_length=100, default="")
    format = models.CharField(verbose_name=_('Format'), max_length=100, choices=REPORT_ITEM_FORMAT_CHOICES)
    report = models.ForeignKey(verbose_name=_('Report'), to=Report, on_delete=models.CASCADE)
    upc = models.CharField(verbose_name=_('UPC/EAN'), max_length=20, blank=True, null=True)
    isrc = models.CharField(verbose_name=_('ISRC'), max_length=20, blank=True, null=True)
    youtube_asset_asset_id = models.CharField(verbose_name=_('Youtube ASSET ID'), max_length=20, blank=True,
                                              null=True)
    dsp_track_id = models.CharField(verbose_name=_('DSP'), max_length=100, default="")
    tax_retention = models.DecimalField(verbose_name=_('Tax Retention'), max_digits=15, decimal_places=5,
                                        default=0.00)
    value = models.DecimalField(verbose_name=_('Value'), max_digits=15, decimal_places=5)
    quantity = models.IntegerField(verbose_name=_('Quantity'), default=0)
    reference = models.TextField(verbose_name=_('Reference'), blank=True)
    country = models.CharField(verbose_name=_('Country'), max_length=30, blank=True)
    product_dsp = models.CharField(verbose_name=_('Product of Use'), max_length=50, blank=True)
    start_date = models.DateField(verbose_name=_('Start Date'), blank=True, null=True)
    end_date = models.DateField(verbose_name=_('End Date'), blank=True, null=True)

    class Meta:
        abstract = True

    def get_title(self):
        return _('N/A')

    def get_product_title(self):
        return _('N/A')

    def get_holder(self):
        return []

    def get_parent_upc(self):
        return ''

    def get_parent_isrc(self):
        return _('N/A')

    def get_artists_names(self):
        names = ''
        try:
            if self.youtube_asset is not None:
                names = Artist.objects.filter(youtube_asset_primary=self.youtube_asset).values_list('name', flat=True)
                # Se nao houverem artistas, retorna o nome do titular
                if not names:
                    return self.youtube_asset.main_holder.name.upper()
        except AttributeError:
            if self.asset is not None:
                names = Artist.objects.filter(asset_primary=self.asset).values_list('name', flat=True)
                # Se nao houverem artistas, retorna o nome do titular
                if not names:
                    return self.asset.main_holder.name.upper()
        return ' + '.join(str(name.upper()) for name in names)

    def get_artists_feat_names(self):
        return ''

    def calculate_values_to_distribute(self, currency_price, holder_share, holder_item_share, provider_share,
                                       tax_retention, transfer_dist_share) -> dict:
        """
            Calcula os valores que serão entregues ao titular, à Oni e ao provider
        Args:
            currency_price: Preço da moeda em Real, usado para conversão
            holder_share: Share do titular (70/30 por ex, 70 é o holder_share)
            holder_item_share: Porcentagem do artista naquele item (porcentagem do holder_share que é dele)
            provider_share: Porcentagem do provedor (Youtube, FUGA, Merlin, Hitbel, etc)
            tax_retention: Retenção de impostos
            transfer_dist_share: Porcentagem do distributor (Oni)

        Returns: Dict com os valores a serem distribuídos neste item para um determinado holder
        """
        # A retenção de imposto é individual de cada holder e deve ser calculada com base na porcentagem que é dele
        tax_retention = tax_retention * holder_item_share / 100
        value_total = decimal.Decimal(self.value * ((100 - provider_share) / 100))
        holder_item_value = holder_item_share * self.value / 100  # Valor total que será distribuído neste holder
        # O valor que é devido ao provedor deve ser calculado encima do que é devido ao holder, para sabermos quanto
        #  da parte do holder deve ir para o provedor
        provider_fee = holder_item_value * (provider_share / 100)  # Valor que é devido ao provedor
        value_total_to_distribute = decimal.Decimal(
            holder_item_value - provider_fee)  # O valor do provedor deve ser reservado

        value_transfer_dist = (decimal.Decimal((100 - holder_share) / 100) * decimal.Decimal(
            transfer_dist_share / 100)) * \
                              decimal.Decimal(value_total)
        #  nesse caso o titular tá devendo dinheiro pá nois
        if holder_item_value == 0 and transfer_dist_share > 0:
            # Parte que será paga ao artista
            value_holder = value_transfer_dist * -1
            value_distributor = value_transfer_dist  # Parte da Oni
        else:
            # Parte que será paga ao artista
            value_holder = round(value_total_to_distribute * decimal.Decimal(holder_share / 100),
                                 5) - value_transfer_dist
            value_distributor = round(value_total_to_distribute - value_holder, 5)  # Parte da Oni

        transfer_dist_share_final = transfer_dist_share * ((100 - holder_share) / 100)

        return {
            'provider_fees': provider_fee,
            'holder_item_share': holder_item_share,
            'value_total': holder_item_value,
            'value_total_to_distribute': value_total_to_distribute,
            'value_holder': value_holder,
            'value_holder_converted': round(value_holder * currency_price, 5),
            'rate_distributor': 100 - holder_share,
            'value_distributor': value_distributor,
            'transfer_dist_share': round(transfer_dist_share_final, 5),
            'value_converted_distributor': round(value_distributor * currency_price, 5),
            'tax_retention': tax_retention,
        }

    def distribute_item(self, report) -> dict:
        """
        Pega os valores e informações de um item e os distribui, criando o registro daquele item no BD com os valores
        individuais corretos. Ao final, acumula os valores individuais deste item ao relatório geral.

        Returns: Dict com os valores do relatório atualizados com os valores deste item
        """
        if report is None:
            report = self.report

        holders = self.get_holder()  # Pega o titular e repasses com valores respectivos
        artists_names = self.get_artists_names()  # Nomes dos artistas principais apenas p registro
        artists_feat_names = self.get_artists_feat_names()  # Nomes dos artistas feat apenas p registro
        report_stats = {
            'gross_revenue_distributed': 0,  # Receita bruta distribuída
            'net_revenue_distributed': 0,  # Receita líquida distribuída (total - share da oni)
            'net_income_distributed': 0,  # Lucro da Oni (tirando o que foi pro titular e pro provider)
            'items_distributed': 0  # Núm de itens distribuídos (um item eh uma linha do relatório)
        }
        for holder in holders:  # Pra cada pessoa que tem que receber alguma grana
            # Calcula os valores a serem distribuídos
            value_to_dist = self.calculate_values_to_distribute(report.currency_price, holder['share'],
                                                                holder['item_share'], report.provider.share,
                                                                self.tax_retention, holder['transfer_dist_share'])
            artists_names = holder['artist_name'] or artists_names  # Dá prioridade ao nome do artista registrado no bd
            # print(value_to_dist)
            item = ReportItem(
                # Dados de identificação
                report_id=report.id,
                holder_id=holder['holder_id'],
                artist_id=holder['artist_id'],
                artists_names=str(artists_names).upper(),
                artists_feat_names=str(artists_feat_names).upper(),
                upc=str(self.get_parent_upc()),
                isrc=str(self.get_parent_isrc()),
                youtube_asset_asset_id=str(self.youtube_asset_asset_id),
                # Dados financeiros
                holder_item_share=value_to_dist['holder_item_share'],  # Percentual do artista sobre esse item
                tax_retention=value_to_dist['tax_retention'],  # Retenção de imposto
                value_holder=value_to_dist['value_holder'],  # Valor que este artista deve receber
                value_holder_converted=value_to_dist['value_holder_converted'],  # Valor devido ao artista na moeda dele
                rate_distributor=value_to_dist['rate_distributor'],  # Percentual da Oni
                value_distributor=value_to_dist['value_distributor'],  # Valor pertencente à Oni
                value_converted_distributor=value_to_dist['value_converted_distributor'],  # Valor da Oni em Real
                transfer_dist_share=value_to_dist['transfer_dist_share'],  # Share extra do distribuidor
                # Dados usados para análises financeiras pós-lançamento
                quantity=self.quantity,
                country=str(self.country),
                product_dsp=str(self.product_dsp),
                dsp_track_id=str(self.dsp_track_id),
                dsp=str(self.dsp),
                format=str(self.format),
                title=str(self.get_title()),
                product_title=str(self.get_product_title()),
                start_date=self.start_date,
                end_date=self.end_date

            )
            item.save()
            try:
                DSPProductCodeTranslation.objects.get(dsp=item.dsp, code=item.product_dsp)
            except DSPProductCodeTranslation.DoesNotExist:
                try:
                    code_translation = DSPProductCodeTranslation(dsp=item.dsp, code=item.product_dsp)
                    code_translation.save()
                except Exception as e:
                    log_error(e)
            # Acumulando valores individuais do item distribuído ao total do relatório
            report_stats['gross_revenue_distributed'] += value_to_dist['value_total']
            report_stats['net_revenue_distributed'] += value_to_dist['value_total_to_distribute']
            report_stats['net_income_distributed'] += value_to_dist['value_distributor']
            report_stats['items_distributed'] += 1

        return report_stats


class ReportTempItemYoutube(ReportTempItem):
    """Mother class to store report items before distribution.
    """

    youtube_asset = models.ForeignKey(verbose_name=_('Youtube Asset'), to=YoutubeAsset, on_delete=models.PROTECT,
                                      blank=True, null=True)

    class Meta:
        verbose_name = _('Report Temp Item Youtube')
        verbose_name_plural = _('Report Temp Items Youtube')

    def get_title(self):
        if self.youtube_asset is not None:
            return self.youtube_asset.title
        else:
            return _('N/A')

    def get_holder(self):
        holders = []
        transfer_share = 0
        # transfer_share_ignore = 0
        transfer_dist_share = 0
        # todo tem que e se não tiver youtube_asset ou main_holder?
        main_holder = self.youtube_asset.main_holder
        for holder in self.youtube_asset.youtubeassetholder_set.all():
            head_share = holder.get_share(main_holder.share_youtube)
            holders.append({
                'holder_id': holder.holder_id,
                'artist_id': holder.artist_id,
                'artist_name': holder.artist.name,
                'item_share': holder.percentage,  # percentage this holder owns of this item
                'share': head_share,  # percentage to be distributed to the holder
                'transfer_dist_share': 0,
            })
            transfer_share += holder.percentage
            transfer_dist_share += holder.percentage if main_holder.share_youtube < head_share else 0

        holders.append({
            'holder_id': main_holder.id,
            'artist_id': None,
            'artist_name': None,
            'item_share': 100 - transfer_share,  # percentage this holder owns of this item
            'share': main_holder.share_youtube,  # percentage to be distributed to the holder
            'transfer_dist_share': transfer_dist_share,
        })
        return holders

    def get_parent_upc(self):
        return ''

    def get_parent_isrc(self):
        return self.youtube_asset.isrc

    def save(self, *args, **kwargs):
        # override para calcular a retenção de imposto para US. Gambiarra bruta
        if self.country == 'US':
            self.tax_retention = self.value * decimal.Decimal(0.3)
            self.value -= self.tax_retention
        super(ReportTempItemYoutube, self).save(*args, **kwargs)


class ReportTempItemAsset(ReportTempItem):
    """Mother class to store report items before distribution.
    """

    asset = models.ForeignKey(verbose_name=_('Asset'), to=Asset, on_delete=models.PROTECT,
                              blank=True, null=True)
    product = models.ForeignKey(verbose_name=_('Product'), to=Product, on_delete=models.PROTECT,
                                blank=True, null=True)

    class Meta:
        verbose_name = _('Report Temp Item DSP')
        verbose_name_plural = _('Report Temp Items DSP')

    def get_title(self):
        if self.asset is not None:
            return self.asset.title
        else:
            return _('N/A')

    def get_product_title(self):
        if self.product is not None:
            return self.product.title
        else:
            return ''

    def get_holder(self):
        holders = []

        transfer_share = 0
        main_holder: Holder = None
        # transfer_share_ignore = 0
        transfer_dist_share = 0
        # log_tests(f'transfer_share_ignore: {transfer_share_ignore}')
        if self.asset is not None:
            main_holder = self.asset.main_holder
            for holder in self.asset.assetholder_set.all():
                head_share = holder.get_share(main_holder.share)
                holders.append({
                    # Informações para identificação
                    'holder_id': holder.holder_id,
                    'artist_id': holder.artist_id,
                    'artist_name': holder.artist.name,
                    # Valores percentuais deste repasse
                    'transfer_dist_share': 0,  # Percentual do repasse que irá pro distributor (Oni)
                    'item_share': holder.percentage,  # Porcentagem do repasse (não confundir com share do holder)
                    'share': head_share,  # Porcentagem do titular
                })
                transfer_share += holder.percentage
                transfer_dist_share += holder.percentage if main_holder.share < head_share else 0
                # log_tests(f'transfer_share_ignore: {transfer_share_ignore}')

        elif self.product is not None:
            main_holder = self.product.main_holder
            for holder in self.product.productholder_set.all():
                head_share = holder.get_share(main_holder.share)
                holders.append({
                    # Informações para identificação
                    'holder_id': holder.holder_id,
                    'artist_id': holder.artist_id,
                    'artist_name': holder.artist.name,
                    # Valores percentuais deste repasse
                    'transfer_dist_share': 0,  # Percentual do repasse que irá pro distributor (Oni)
                    'item_share': holder.percentage,  # Porcentagem do repasse (não confundir com share do holder)
                    'share': head_share,  # Porcentagem do titular
                })
                transfer_share += holder.percentage
                transfer_dist_share += holder.percentage if main_holder.share < head_share else 0

        holders.append({
            # Informações para identificação
            'holder_id': main_holder.id,
            'artist_id': None,
            'artist_name': None,
            # Valores percentuais deste repasse
            'transfer_dist_share': transfer_dist_share,
            'item_share': 100 - transfer_share,  # O percentual do principal é o que sobrou dos repasses
            'share': main_holder.share,  # Porcentagem do titular
        })
        return holders

    def get_parent_upc(self):
        try:
            return self.product.upc
        except AttributeError:
            return ''

    def get_parent_isrc(self):
        try:
            return self.asset.isrc
        except AttributeError:
            return ''

    def get_artists_feat_names(self):
        if self.asset is not None:
            names = Artist.objects.filter(asset_feat=self.asset).values_list('name', flat=True)
            return ','.join(str(name.upper()) for name in names)

        return ''


class ReportItem(BaseModel):
    """ReportItem is the final distributed item
    """

    report = models.ForeignKey(verbose_name=_('Report'), to=Report, on_delete=models.CASCADE)
    holder = models.ForeignKey(verbose_name=_('Holder'), to=Holder, on_delete=models.PROTECT)
    artist = models.ForeignKey(verbose_name=_('Artist'), to=Artist, on_delete=models.PROTECT, blank=True, null=True)

    artists_names = models.TextField(verbose_name=_('Artists'), default='')
    artists_feat_names = models.TextField(verbose_name=_('Artists Feat.'), default='')

    upc = models.CharField(verbose_name=_('UPC/EAN'), max_length=20, blank=True)
    isrc = models.CharField(verbose_name=_('ISRC'), max_length=20, blank=True)
    youtube_asset_asset_id = models.CharField(verbose_name=_('Youtube Asset ID'), max_length=30, blank=True)

    holder_item_share = models.DecimalField(verbose_name=_('Holder Item Share'), max_digits=15, decimal_places=5,
                                            default=0.00)

    tax_retention = models.DecimalField(verbose_name=_('Tax Retention'), max_digits=15, decimal_places=5,
                                        default=0.00)

    value_holder = models.DecimalField(verbose_name=_('Value Holder'), max_digits=15, decimal_places=5)
    value_holder_converted = models.DecimalField(verbose_name=_('Value Holder Converted'), max_digits=15,
                                                 decimal_places=5)

    rate_distributor = models.DecimalField(verbose_name=_('Rate Dist.'), max_digits=15, decimal_places=5)
    value_distributor = models.DecimalField(verbose_name=_('Value Dist.'), max_digits=15, decimal_places=5)
    value_converted_distributor = models.DecimalField(verbose_name=_('Value Dist. Converted'), max_digits=15,
                                                      decimal_places=5)
    transfer_dist_share = models.DecimalField(verbose_name=_('Extra Dist. Share'), max_digits=15, decimal_places=5,
                                              default=0.00)

    quantity = models.IntegerField(verbose_name=_('Quantity'), default=0)

    country = models.CharField(verbose_name=_('Country'), max_length=30, blank=True)
    product_dsp = models.CharField(verbose_name=_('Product'), max_length=50, blank=True)
    dsp = models.CharField(verbose_name=_('DSP'), max_length=100, default="")
    dsp_track_id = models.CharField(verbose_name=_('DSP Track ID'), max_length=100, default="")

    format = models.CharField(verbose_name=_('Format'), max_length=100, choices=REPORT_ITEM_FORMAT_CHOICES)
    title = models.CharField(verbose_name=_('Title'), max_length=100)
    product_title = models.CharField(verbose_name=_('Album'), max_length=100)

    start_date = models.DateField(verbose_name=_('Start Date'), blank=True, null=True)
    end_date = models.DateField(verbose_name=_('End Date'), blank=True, null=True)
    backup_done = models.BooleanField(default=False, verbose_name=_('Backup Done'))

    class Meta:
        verbose_name = _('Report Item')
        verbose_name_plural = _('Report Items')

    def __str__(self):
        """str method"""
        return "{} ({})".format(self.title, self.dsp)

    @property
    def total_value(self):
        return self.value_holder + self.value_distributor

    @property
    def total_value_converted(self):
        return self.value_holder_converted + self.value_converted_distributor

    @property
    def share_holder(self):
        return decimal.Decimal(100.00) - self.rate_distributor

    @property
    def unit_price(self) -> decimal:
        try:
            return round(self.total_value / self.quantity, 6)
        except (ZeroDivisionError, decimal.InvalidOperation):
            return 0.00

    @property
    def gross_total_value(self) -> decimal:
        #  total global recebido por todos titulares da obra/produto/asset
        try:
            return round((100 * self.total_value) / self.holder_item_share, 2)
        except (ZeroDivisionError, decimal.InvalidOperation):
            return 0.00

    @property
    def gross_unit_price(self) -> decimal:
        try:
            return round(self.gross_total_value / self.quantity, 2)
        except (ZeroDivisionError, decimal.InvalidOperation):
            return 0.00


def get_invoice_files_paths(instance, filename, filename_custom=''):
    import os
    ext = filename.split('.')[-1]
    batch_name = instance.batch.name
    filename_new = '{} - {}{}.{}'.format(batch_name, instance.holder.name, filename_custom, ext)
    return os.path.join('reports', batch_name, filename_new)


def get_invoice_pdf_file_path(instance, filename):
    return get_invoice_files_paths(instance, filename)


def get_invoice_items_file_path(instance, filename):
    return get_invoice_files_paths(instance, filename)


def get_invoice_summary_file_path(instance, filename):
    filename_custom = ' {}'.format(_('Summary'))
    return get_invoice_files_paths(instance, filename, filename_custom)


def get_invoice_distributor_file_path(instance, filename):
    filename_custom = ' {}'.format(_('Invoice Dist.'))
    return get_invoice_files_paths(instance, filename, filename_custom)


class Invoice(BaseModel):
    """Invoice is a class used to lock a total value """
    holder = models.ForeignKey(verbose_name=_('Holder'), to=Holder, on_delete=models.PROTECT)
    batch = models.ForeignKey(verbose_name=_('Batch'), to=Batch, on_delete=models.CASCADE)

    value_distributor_converted = models.DecimalField(verbose_name=_('Value Dist.'), max_digits=15,
                                                      decimal_places=2)
    value_converted = models.DecimalField(verbose_name=_('Value'), max_digits=15, decimal_places=2)
    value_commission_converted = models.DecimalField(verbose_name=_('Value Comission'), max_digits=15,
                                                     decimal_places=2,
                                                     null=True, blank=True)
    tax_retention = models.DecimalField(verbose_name=_('Tax Retention'), decimal_places=2, max_digits=15,
                                        default=0.00)

    payment_date = models.DateField(verbose_name=_('Payment Date'), null=True, blank=True)
    number = models.CharField(verbose_name=_('Number'), max_length=100, unique=True)
    status = models.CharField(verbose_name=_('Status'), max_length=3, choices=INVOICE_STATUS_CHOICES, default='OP')
    bank_data = models.TextField(verbose_name=_('Bank Data'), blank=True)
    notes = models.TextField(verbose_name=_('Notes'), blank=True)

    legal_name = models.CharField(verbose_name=_('Legal Name'), max_length=100, blank=True)
    legal_document = models.CharField(verbose_name=_('Legal Document'), max_length=100)

    file_items = models.FileField(verbose_name=_('Items Details'), upload_to=get_invoice_items_file_path,
                                  max_length=250,
                                  null=True, blank=True)
    file_pdf = models.FileField(verbose_name=_('Invoice'), upload_to=get_invoice_pdf_file_path, max_length=250,
                                null=True, blank=True)
    file_summary = models.FileField(verbose_name=_('Summary'), upload_to=get_invoice_summary_file_path,
                                    max_length=250,
                                    null=True, blank=True)
    file_invoice_distributor = models.FileField(verbose_name=_('Invoice Dist.'),
                                                upload_to=get_invoice_distributor_file_path,
                                                max_length=250,
                                                null=True, blank=True)
    commission = models.DecimalField(verbose_name=_('Commission'), decimal_places=3, max_digits=6,
                                     validators=[MaxValueValidator(100), MinValueValidator(0)], default=0)
    processing_percentage = models.FloatField(verbose_name=_('% Processing'), default=0.0)

    BULK_FIELDS = (
        ('numero', _('REQUIRED. Used to identify the invoice on the batch.')),
        ('data_pgto', _('OPTIONAL. Format dd/mm/aaaa.')),
        ('ret_imposto',
         _('OPTIONAL. Use period instead of comma and dont use %. Eg.: 1.5 (for 1,5%).')),
        ('comissao', _('OPTIONAL. Use period instead of comma and dont use %. Eg.: 1.5 (for 1,5%).')),
        ('notas', _('OPTIONAL. Text only.')),
    )

    @property
    def federal_retention(self) -> decimal.Decimal:
        """
        Calcula retenção de imposto para a ND.
        """
        if 'CNPJ' not in self.holder.legal_document:  # R.N.: Pessoa física não retém imposto
            return decimal.Decimal(0.00)
        value = self.value_distributor_converted * decimal.Decimal(0.015)
        return round(value, 2) if value > decimal.Decimal(10.00) else decimal.Decimal(0.00)

    @property
    def to_pay(self):
        return round(decimal.Decimal(self.value_converted) + decimal.Decimal(self.federal_retention) - decimal.Decimal(
            self.tax_retention), 2)

    @property
    def to_deposit(self):
        return self.to_pay - self.advance_value_total

    @property
    def payment_date_for_humans(self):
        try:
            return self.payment_date.strftime('%d/%m/%Y')
        except Exception:
            return _('N/A')

    @property
    def net_distributed_value(self):
        # Valor do distribuidor menos comissão para selo.
        return self.value_distributor_converted - self.value_commission_converted if self.value_commission_converted else self.value_distributor_converted

    net_distributed_value.fget.short_description = _('Net Val. Dist')

    class Meta:
        verbose_name = _('Invoice')
        verbose_name_plural = _('Invoices')
        ordering = ['-id', ]
        permissions = [('can_set_invoice_as_paid', _('Can set invoice as paid'))]

    def __str__(self):
        """str method"""
        return f"{self.holder.name} ({self.value_converted})"

    def mark_as_paid(self):
        """Marca a invoice como paga"""
        self.status = 'CL'
        self.payment_date = self.batch.invoice_due_date
        self.save()

    @property
    def advance_value_total(self):
        """ Retorna o valor do adiantamento
        """
        return self.advanceinvoicerelation_set.aggregate(total=Sum('amount')).get('total', 0) or 0

    def get_data_for_api(self):
        return {
            'id': self.id,
            'holder': self.holder_id,
            'batch': self.batch_id,
            'value_distributor_converted': self.value_distributor_converted,
            'value_converted': self.value_converted,
            'value_commission_converted': self.value_commission_converted,
            'tax_retention': self.tax_retention,
            'net_distributed_value': self.net_distributed_value,
            'payment_date': self.payment_date_for_humans,
            'number': self.number,
            'status': self.status,
            'bank_data': self.bank_data,
            'notes': self.notes,
            'legal_name': self.legal_name,
            'legal_document': self.legal_document,
            'file_items': self.file_items.name,
            'file_pdf': self.file_pdf.name,
            'file_summary': self.file_summary.name,
            'file_invoice_distributor': self.file_invoice_distributor.name,
            'commission': self.commission,
            'processing_percentage': self.processing_percentage,
            'batch_name': self.batch.__str__(),
            'holder_name': self.holder.__str__(),
            'get_status_display': self.get_status_display(),
        }

    def make_files(self):
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        from io import BytesIO as IO
        from openpyxl import Workbook

        # processamento 0% completo
        self.processing_percentage = 0.00
        self.save()

        pdf_temp = IO()
        template = get_template("label_reports/invoice_pdf.html")
        context_pdf = {

            'invoice_number': self.number,
            'invoice_created_date': self.created_at,
            'invoice_due_date': self.batch.invoice_due_date,
            'invoice_from_name': self.holder.legal_name if self.holder.legal_name != '' else '-',
            'invoice_from_document': self.holder.legal_document if self.holder.legal_document != '' else '-',
            'invoice_logo_url': self.holder.get_master_client_logo_url(),
            'invoice_to_name': INVOICE_SETTINGS['COMPANY_NAME'],
            'invoice_to_document': INVOICE_SETTINGS['COMPANY_DOCUMENT'],
            'invoice_to_extra': INVOICE_SETTINGS['COMPANY_ADDRESS'],
            'invoice_product': INVOICE_SETTINGS['PRODUCT_DESCRIPTION'],
            'invoice_bank_data': self.holder.bank_data,
            'invoice_total': self.gross_value_converted,
            'invoice_tax_retention': self.tax_retention,
            'invoice_net_total': self.get_net_value,
            'advance_total': self.advance_value_total,
        }
        html_pdf = template.render(context=context_pdf)
        pisa.pisaDocument(src=IO(html_pdf.encode("UTF-8")), dest=pdf_temp, name='')
        self.file_pdf.save(name='pdf', content=pdf_temp)

        # processamento 15% completo
        self.processing_percentage = 15.00
        self.save()

        summary_temp = IO()
        template = get_template("label_reports/summary_pdf.html")

        artists = self.get_value_holder_converted_by_artists()
        dsps = self.get_value_holder_converted_by_dsps()

        context_summary = {
            'invoice_total': self.gross_value_converted,
            'invoice_from_name': self.holder.name,
            'invoice_number': self.number,
            'invoice_created_date': self.created_at,
            'invoice_due_date': self.batch.invoice_due_date,
            'invoice_logo_url': self.holder.get_master_client_logo_url(),
            'artists': artists,
            'dsps': dsps,
        }
        html_summary = template.render(context=context_summary)
        pisa.pisaDocument(src=IO(html_summary.encode("UTF-8")), dest=summary_temp, name='')
        self.file_summary.save(name='pdf', content=summary_temp)

        # processamento 30% completo
        self.processing_percentage = 30.00
        self.save()

        # excel_file = IO()
        # report_items_list = []
        report_items = ReportItem.objects.filter(holder_id=self.holder_id, report__batch=self.batch).order_by(
            '-value_holder_converted').prefetch_related('holder').prefetch_related('artist')

        # processamento 60% completo
        self.processing_percentage = 60.00
        self.save()

        # batch_base_currency = self.batch.base_currency
        report_currencies = {}
        for report in Report.objects.filter(batch=self.batch):
            report_currencies[report.id] = {
                'currency': report.currency,
                'price': report.currency_price,
            }

        # processamento 70% completo
        self.processing_percentage = 70.00
        self.save()

        # artist_names = {}
        # for artist in Artist.objects.all():
        #     artist_names[artist.id] = artist.name
        holder_names = {}
        for holder in Holder.objects.all():
            holder_names[holder.id] = holder.name

        # processamento 80% completo
        self.processing_percentage = 80.00
        self.save()

        workbook = Workbook()
        worksheet = workbook.create_sheet('data')
        workbook.remove(workbook.get_sheet_by_name('Sheet'))
        # font_style = xlwt.XFStyle()
        # font_style.font.bold = True
        titles = [str(_('Start Date')),
                  str(_('End Date')),
                  str(_('Holder')),
                  str(_('Title')),
                  str(_('Album')),
                  str(_('Artists')),
                  str(_('Artists Feat.')),
                  str(_('UPC')),
                  str(_('ISRC')),
                  str(_('Youtube Asset ID')),
                  str(_('Product')),
                  str(_('DSP')),
                  str(_('DSP Track ID')),
                  str(_('Format')),
                  str(_('Country')),

                  str(_('Quantity')),

                  str(_('Unit Gross Price')),
                  str(_('Gross Total')),

                  str(_('Holder Item Share')),
                  str(_('Unit Net Price')),
                  str(_('Net Total')),

                  str(_('Currency')),
                  str(_('Currency Price')),
                  str(_('Net Total Converted')),

                  str(_('Share Dist.')),
                  str(_('Extra Dist. Share')),
                  str(_('Value Dist.')),
                  str(_('Value Dist. Converted')),

                  str(_('Share Holder')),
                  str(_('Value Holder')),
                  str(_('Value Holder Converted'))]
        worksheet.append(titles)
        for cell in range(len(titles)):
            worksheet.cell(1, cell + 1).style = 'Headline 3'

        total_items = len(report_items)
        rate = int(total_items / 10)

        for item_number, report_item in enumerate(report_items, start=1):

            # de tempos em tempos, incrementaremos a percentagem de processamento de acordo com quantas linhas ja foram
            try:
                if item_number % rate == 0:
                    self.processing_percentage += (item_number / total_items) * 20
                    self.save()
            except Exception:
                # teremos problemas quando esse resultado rate = int(total_items / 10) for 0
                pass

            worksheet.append([
                report_item.start_date,
                report_item.end_date,
                holder_names[report_item.holder_id],  #

                report_item.title,
                report_item.product_title,
                report_item.artists_names,  #
                report_item.artists_feat_names,  #
                report_item.upc,
                report_item.isrc,
                report_item.youtube_asset_asset_id,

                report_item.product_dsp,
                report_item.dsp,
                report_item.dsp_track_id,
                report_item.format,
                report_item.country,

                report_item.quantity,
                report_item.gross_unit_price,
                report_item.gross_total_value,

                report_item.holder_item_share,
                report_item.unit_price,
                report_item.total_value,

                report_currencies[report_item.report_id]['currency'],  #
                report_currencies[report_item.report_id]['price'],  #
                report_item.total_value_converted,
                # report_item.holder_item_share,

                report_item.rate_distributor,
                report_item.transfer_dist_share,
                report_item.value_distributor,
                report_item.value_converted_distributor,

                report_item.share_holder,
                report_item.value_holder,
                report_item.value_holder_converted,

            ])

        excel_file = IO()

        workbook.save(excel_file)
        self.file_items.save(name='excel.xlsx', content=excel_file)

        # processamento 100% completo
        self.processing_percentage = 100.00
        self.save()

    @staticmethod
    def get_closed_status():
        return 'CL'

    @property
    def tax_retention_for_humans(self):
        try:
            currency_symbol = CURRENCY_SYMBOLS[self.batch.base_currency]
        except KeyError:
            currency_symbol = '$'
        return '{} {}'.format(currency_symbol, self.tax_retention)

    @property
    def net_value_converted_for_humans(self):
        try:
            currency_symbol = CURRENCY_SYMBOLS[self.batch.base_currency]
        except KeyError:
            currency_symbol = '$'
        return '{} {}'.format(currency_symbol, self.value_converted)

    @property
    def gross_value_converted(self):
        return round(self.value_converted, 2)

    @property
    def get_net_value(self):
        return round(self.value_converted - self.tax_retention - self.advance_value_total, 2)

    @property
    def value_converted_gross_for_humans(self):
        try:
            currency_symbol = CURRENCY_SYMBOLS[self.batch.base_currency]
        except KeyError:
            currency_symbol = '$'
        return '{} {}'.format(currency_symbol, self.gross_value_converted)

    def get_value_holder_converted_by_artists(self) -> []:
        return ReportItem.objects.values(name=F('artists_names')).annotate(
            value_converted=Sum('value_holder_converted'),
            quantity=Sum('quantity')).filter(
            holder_id=self.holder_id, report__batch=self.batch).order_by('-value_converted')

    def get_value_holder_converted_by_dsps(self) -> []:
        return ReportItem.objects.values(name=F('dsp')).annotate(value_converted=Sum('value_holder_converted'),
                                                                 quantity=Sum('quantity')).filter(
            holder_id=self.holder_id, report__batch=self.batch).order_by('-value_converted')

    def get_can_make_invoice_file(self):
        """Split and concatenates all errors messages
        """
        button = "<a href=\"{}\" class=\"default\">{}</a>".format(
            reverse('label_reports:reports.make_invoice_file', args=[self.id]),
            _('Make Files'))
        return format_html('<span class="icon-tick"></span>' + button)

    get_can_make_invoice_file.short_description = _('Make Invoices')

    @staticmethod
    def query_invoices_by_args(request) -> dict:
        """
            Metodo usado pela api do DataTables para buscar dinamicamente por invoices com base na caixa de busca
            Args:
                request: request da api
            Returns:
                dict contendo a queryset de invoices e outras informacoes relevantes ao DataTables
        """
        return default_query_assets_by_args(request, Invoice)

    @staticmethod
    def mark_as_paid_bulk(queryset):
        from django.db.models import OuterRef, Subquery
        queryset.update(
            status='CL',
            payment_date=Subquery(
                Batch.objects.filter(
                    id=OuterRef('batch_id')
                ).values('invoice_due_date')[:1]
            )
        )

    @staticmethod
    def get_column_order_choices() -> List[str]:
        """Retorna o dicionario com as colunas do datatables em que os produtos podem ser ordenados"""
        return INVOICE_ORDER_COLUMN_CHOICES

    @classmethod
    def filter_objects(cls, searched_value: str, request_user: User, values_list_fields: list = None,
                       custom_query: Q = Q(), initial_queryset: QuerySet('Invoice') = None) -> QuerySet:
        """Filtra os objetos da classe com base na string passada como parâmetro. Vide docstring do método chamado."""
        queryset = initial_queryset or cls.objects.all()
        search_fields = ['holder__name', 'batch__name', 'number']
        # Caso o usuario tenha passado uma query como parametro, o filtro sera feito com base nela apenas
        if custom_query:
            queryset = queryset.filter(custom_query)
        return ObjectFilterer.filter_objects(cls, searched_value, request_user.user_user_profile, search_fields,
                                             queryset, values_list_fields)


class DSPProductCodeTranslation(BaseModel):
    code = models.CharField(max_length=50, verbose_name=_('Report Code'),
                            help_text=_('This is the code that identifies the product type on the report.'))
    dsp = models.CharField(verbose_name=_('DSP'), max_length=100, default="")
    translation = models.CharField(max_length=250, verbose_name=_('Translation'),
                                   help_text=_('This is the translation for the code that comes in the report.'),
                                   null=True, blank=True)

    class Meta:
        verbose_name = _('DSP Product Code Translation')
        verbose_name_plural = _('DSP Product Code Translations')
        permissions = [('can_backup_code_translations', _('Can Backup Code Translations'))]

    def __str__(self):
        return f'({self.dsp.__str__()}){self.code}'

    def save(self, force_insert=False, force_update=False, using=None, update_fields=None):
        ids = [self.id] if self.id else []
        if not DSPProductCodeTranslation.objects.filter(code=self.code, dsp=self.dsp).exclude(id__in=ids).exists():
            super(DSPProductCodeTranslation, self).save(force_insert, force_update, using, update_fields)

    @staticmethod
    def bigquery_table_schema():
        """Table Schema para backup no BigQuery"""
        return [
            {'name': 'code', 'type': 'STRING'},
            {'name': 'dsp', 'type': 'STRING'},
            {'name': 'translation', 'type': 'STRING'},
        ]

    @staticmethod
    def bigquery_dataframe_data_types():
        """DataTypes para backup no BigQuery"""
        return {
            'code': 'string',
            'dsp': 'string',
            'translation': 'string',
        }

    @classmethod
    def perform_backup(cls) -> bool:
        """Realiza o backup deste modelo para o Google"""
        fields = cls.get_model_fields(exclude=['id', 'created_at', 'updated_at'])
        queryset = cls.objects.all()
        data = [[obj.code, obj.dsp, obj.translation] for obj in queryset]
        bqb_handler = BackupHandler()
        # Limpa o conteúdo da tabelas pra atualizar dados em vez de replicá-los
        bqb_handler.clear_table_content()
        # Faz o backup dos dados para o google storage e big query
        return bqb_handler.perform_backup(fields, data, GOOGLE_STORAGE_DATASET_ID, GOOGLE_STORAGE_BACKUP_BUCKET,
                                          'traducoes_cod_produtos_dsp',
                                          f'traducoes_cod_produtos_dsp/dsp_product_codes_{timezone.now().strftime("%d_%m_%Y")}',
                                          GOOGLE_STORAGE_PARENT, 'traducoes_cod_produtos_dsp', queryset.count())


class Advance(BaseModel):
    """ Classe reponsável por gerenciar adiantamentos no App Oni
    """

    class DiscountTypeChoices(models.TextChoices):
        INSTALLMENTS = 'INST', _('Installments')
        PERCENTAGE = 'PERC', _('Percentage')

    class AdvanceTypeChoices(models.TextChoices):
        ADVANCE = 'ADV', _('Advance')
        LOAN = 'LN', _('Loan')
        SLEEVE = 'SLV', _('Sleeve')

    holder = models.ForeignKey(to=Holder, on_delete=models.PROTECT, verbose_name=_('Holder'))
    amount = models.DecimalField(max_digits=12, decimal_places=2, verbose_name=_('Amount'))
    advance_date = models.DateField(verbose_name=_('Advance Date'), help_text=_('Date at which the advance was made'))
    payment_start_date = models.DateField(verbose_name=_('Payment Start Date'),
                                          help_text=_('Date at which the payments will start'))
    advance_type = models.CharField(max_length=3, choices=AdvanceTypeChoices.choices, verbose_name=_('Advance Type'),
                                    default=AdvanceTypeChoices.ADVANCE)
    discount_type = models.CharField(max_length=4, choices=DiscountTypeChoices.choices, verbose_name=_('Discount Type'),
                                     default=DiscountTypeChoices.INSTALLMENTS)
    discount_amount = models.DecimalField(max_digits=12, decimal_places=2, verbose_name=_('Discount Amount'),
                                          help_text=_('If the discount type is percentage, max value is 100'))
    will_be_discounted = models.BooleanField(verbose_name=_('Will Be Discounted'), default=True, help_text=_(
        'Indicates whether or not the advance will be discounted in the next invoice'))
    description = models.TextField(verbose_name=_('Description'), null=True, blank=True)

    class Meta:
        verbose_name = _('Advance')
        verbose_name_plural = _('Advances')
        ordering = ['advance_date', 'id']

    def __str__(self):
        return f'{self.get_advance_type_display().upper()} - {self.holder} - {self.amount} ({self.advance_date.strftime("%d/%m/%Y")})'

    def save(self, force_insert=False, force_update=False, using=None, update_fields=None):
        """ Sobrescrevendo o método save para atualziar o valor das amorizações em aberto toda vez que o adiantamento
            for alterado. Isso é feito pra manter a consistência. Caso um adiantamento seja alterado depois que a ND
            foi gerada, essa alteração tem que refletir na hora de gerar novamente o arquivo dessa ND. Atualizando o
            valor das amortizações em aberto, essa regra de negócio é atendida.
        """
        if self.is_fully_paid:
            raise ValueError(_('Paid advances cannot be edited'))
        super(Advance, self).save()
        open_installments = self.advanceinvoicerelation_set.filter(Q(paid=False), Q(Q(invoice__isnull=True) | Q(
            invoice__holder=self.holder))).distinct()  # O filtro tem que pegar as invoices manuais e as de ND.
        for open_installment in open_installments:
            open_installment.set_value()
            open_installment.save()

    @property
    def has_to_be_discounted(self):
        """ Booleano que indica se o adiantamento deve ser descontado na próxima invoice
        """
        if not self.will_be_discounted:
            return False
        if self.is_fully_paid:
            return False
        if self.payment_start_date < timezone.now().date():
            return False
        return True

    def create_advance_invoice_relation(self, invoice: Invoice, max_installment_value: decimal.Decimal):
        """ Cria um objeto do tipo AdvanceInvoiceRelation para representar uma 'parcela' do adiantamento
        Args:
            invoice: nota de débito relacionada à amortização que vai ser criada
            max_installment_value: valor máximo permitido para a amortização (para os casos de mais de um adiantamento)
        """
        if not self.has_to_be_discounted:
            return None
        try:
            installment = AdvanceInvoiceRelation.objects.get(advance=self, invoice=invoice)
        except AdvanceInvoiceRelation.DoesNotExist:
            installment = AdvanceInvoiceRelation(advance=self, invoice=invoice)
        installment.set_value(max_installment_value)
        installment.save()
        return installment

    @property
    def total_paid(self):
        """ Retorna o valor total pago pelo adiantamento
        """
        return self.advanceinvoicerelation_set.filter(paid=True).aggregate(total_paid=models.Sum('amount')).get(
            'total_paid', 0) or 0

    @property
    def total_to_pay(self):
        """ Retorna o valor a ser pago
        """
        return self.amount - self.total_paid

    @property
    def is_fully_paid(self):
        """ Booleano que indica se o adiantamento está totalmente pago
        """
        return self.total_paid >= self.amount

    def next_installment_value(self, amount):
        """ Retorna o valor do próximo desconto. Garante que o valor das parcelas não ultrapasse o valor do adiantamento
        """
        return self.total_to_pay if amount > self.total_to_pay else amount

    @classmethod
    def filter_objects(cls, searched_value: str, request_user: User, values_list_fields: list = None,
                       custom_query: Q = Q(), initial_queryset: QuerySet('Advance') = None) -> QuerySet:
        """Filtra os objetos da classe com base na string passada como parâmetro. Vide docstring do método chamado."""
        queryset = initial_queryset or cls.objects.all()
        search_fields = ['holder__name', 'amount', 'advance_date', 'payment_start_date', 'discount_amount']
        # Caso o usuario tenha passado uma query como parametro, o filtro sera feito com base nela apenas
        if custom_query:
            queryset = queryset.filter(custom_query)
        return ObjectFilterer.filter_objects(cls, searched_value, request_user.user_user_profile, search_fields,
                                             queryset, values_list_fields)


class AdvanceInvoiceRelation(BaseModel):
    """ Classe reponsável por gerenciar relação entre adiantamentos e relatórios
    """

    advance = models.ForeignKey(to=Advance, on_delete=models.PROTECT, verbose_name=_('Advance'))
    invoice = models.ForeignKey(to=Invoice, on_delete=models.PROTECT, verbose_name=_('Invoice'), null=True, blank=True)
    amount = models.DecimalField(max_digits=12, decimal_places=2, verbose_name=_('Amount'))
    paid = models.BooleanField(verbose_name=_('Paid'), default=False)

    class Meta:
        verbose_name = _('Amortization')
        verbose_name_plural = _('Amortizations')

    def __str__(self):
        return f'{self.advance.holder} - {self.advance.advance_date.strftime("%d/%m/%Y")} (R$ {self.amount}/ R${self.advance.amount}){" - PG" if self.paid else ""}'

    def set_value(self, invoice_value: decimal.Decimal = None):
        """ Atualiza o valor da parcela
        Args:
            invoice_value: valor máximo permitido para a amortização. Quano há mais de um adiantamento sendo descontado,
                           esse valor garante que o valor da segunda amortização em diante não ultrapasse o valor da ND.
        """
        if not self.invoice:  # Guard clause. Nesse caso, é uma amortização manual. Nada da lógica abaixo se aplica.
            return self.amount
        advance = self.advance
        invoice_value = invoice_value or self.invoice.value_converted
        if advance.discount_type == advance.DiscountTypeChoices.INSTALLMENTS:
            installment_value = advance.next_installment_value(advance.discount_amount)
        elif advance.discount_type == advance.DiscountTypeChoices.PERCENTAGE:
            installment_value = advance.next_installment_value(round(
                decimal.Decimal(invoice_value) * decimal.Decimal(advance.discount_amount / 100), 2))
        else:
            raise ValueError('Invalid discount type')
        if installment_value > invoice_value:  # Garante que o desconto não ultrapasse o valor da ND
            installment_value = invoice_value
        self.amount = installment_value
        return self.amount


# noinspection PyUnusedLocal
@receiver(pre_delete, sender=Report)
def protect_reports(sender, instance, using, **kwargs):
    """Do not allow reports that are processing to be deleted """
    if not Report.can_delete_report_any(instance):
        raise Exception('Only processed reports can be deleted.')


# ATENCAO!!!! O post-save abaixo foi retirado pq estava sendo inseguro usar este sinal pra executar a tarefa de
# validar o relatorio apos o seu salvamento. O que ocorria é que o sinal de post-save era disparado antes de o objeto
# chegar no BD propriamente, devido a atrasos no bucket (aws). Isto provocava um ObjectDoesNotExist na tarefa, mesmo
# o id ja existindo. Com isso, o disparo desta tarefa foi transferido para um botao no admin de Report.
# @receiver(post_save, sender=Report)
# def validate_reports(sender, instance, using, **kwargs):
#     """If save represents a creation. Dispatch validation task. """
#     if kwargs.get('created', False):
#         from ..tasks import validate_report_task
#         validate_report_task.apply_async((instance.id,),
#                                          eta=timezone.now() + timezone.timedelta(seconds=50))


# @receiver(pre_save, sender=Batch)
# def protect_batch(sender, instance, using, **kwargs):
#     """Do not allow reports that are processing to be deleted
#     """
#     if not Batch.can_change_batch_any(instance):
#         raise Exception('Only open Batches can be updated.')
#     pass


# @receiver(pre_save, sender=Invoice)
# def protect_invoices(sender, instance, using, **kwargs):
#     """Do not allow reports that are processing to be deleted
#     """
#     # todo regra de negocio
#     if instance.status == Invoice.get_closed_status():
#         raise Exception('Only open Invoices can be updated.')
#     pass

@receiver(post_save, sender=AdvanceInvoiceRelation)
def post_save_amortization(sender, instance: AdvanceInvoiceRelation, using, **kwargs):
    """ Para amortizações manuais, marca-as como paga assim que elas são lançadas no sistema
    """
    if kwargs.get('created', False) and not instance.invoice:
        instance.paid = True
        instance.save()


# noinspection PyUnusedLocal
@receiver(post_save, sender=Invoice)
def post_save_invoices(sender, instance: Invoice, using, **kwargs):
    """Invoice post save signal handler

    Mark invoice as payed after changing the pay date. Also mark batch as payed and set correct number.
    Sends notifications
    """
    if kwargs.get('created', False):
        instance.commission = (instance.holder.share / 100 *
                               instance.holder.catalog.sales_commission / 100) * 100
        instance.value_commission_converted = round(
            (instance.value_converted * instance.holder.catalog.sales_commission) / 100, 4)

        # batch__invoice_due_date = instance.batch.invoice_due_date if instance.batch.invoice_due_date is not None else _(
        #     'n/a')
        # send_notification(author=instance, recipients=User.objects.filter(holderuser__holder_id=instance.holder.id),
        #                   verb=_('was processed'), target=instance,
        #                   url=reverse('artists:artists.invoice.show', args=[instance.id]), send_email=True,
        #                   email_template='info',
        #                   email_description='{} {} {} {} {}'.format(_('Invoice'), instance.batch.name,
        #                                                             _('with due date for'),
        #                                                             time.strftime('%d/%m/%Y',
        #                                                                           batch__invoice_due_date.timetuple()),
        #                                                             _('has been processed')))

        instance.number = f'{INVOICE_SETTINGS["NUMBER_PREFIX"]} {str(instance.id)}'
        instance.save()
    if instance.payment_date is not None:
        batch = Batch.objects.get(id=instance.batch_id)
        if instance.status != Invoice.get_closed_status():
            instance.status = Invoice.get_closed_status()
            instance.save()
        if advances := instance.advanceinvoicerelation_set.all():  # marca os adiantamentos como pagos
            advances.update(paid=True)

        if batch.status != 'PA' and instance.status == Invoice.get_closed_status():
            batch.status = 'PA'
            batch.save()
            instance.save()

    if instance.status == Invoice.get_closed_status() and instance.payment_date is not None:
        author = instance
        recipients = User.objects.filter(holderuser__holder_id=instance.holder.id)
        verb = _('approved label')
        action_object = instance
        url = f"{reverse('artists:artists.labels')}{instance.id}"
        email_url = '{}{}'.format('SITE_URL', url)
        if len(recipients) > 0:
            email_logo = recipients[0].user_user_profile.get_master_client_email_logo_url()
            try:
                email_master_client_name = recipients[0].user_user_profile.get_master_client().name
            except AttributeError:
                email_master_client_name = 'FRONT_END__SITE_NAME'
            if author is None:
                author = recipients[0].user_user_profile.get_default_system_master_client()
                # No caso extremo de não haver um master client no sistema, colocamos um autor qualquer
                if not author:
                    log_error('Não há um master client no sistema. Favor corrigir.')
                    author = recipients[0]
            email_description = f'{author} - {verb}: {action_object}' if action_object else f'{author} - {verb}'
            # bell notification
            notify.send(sender=author, recipient=recipients, verb=verb, action_object=action_object, url=url,
                        emailed=True, level='info')
            # todo quando o ator da notificação for um usuário, colocar o nome dele como ator pra melhorar a legibilidade

            # email notification management
            email_support = _('Any questions? Email us!')
            email_support_mail = 'SUPPORT_MAIL'
            email_site_name = 'FRONT_END__SITE_NAME'
            context = {
                'url': email_url,
                'email_title': email_site_name,
                'email_subject': f'{email_site_name}',
                'email_description': email_description,
                'email_button_text': _('Go'),
                'email_support': email_support,
                'email_support_mail': email_support_mail,
                'email_site_name': email_site_name,
                'publisher_logo_path': email_site_name,
                'email_logo': email_logo,
                'email_master_client_name': email_master_client_name,
            }
            email_recipients = []
            for recipient in recipients:
                email_recipients.append(recipient.email)

                if recipient.email is not None and recipient.email != '':
                    email_recipients.append(recipient.email)
            try:
                mail.send(
                    email_recipients,
                    # subject=email_subject,
                    template='level',
                    context=context,
                )
            except ValidationError as e:
                log_error(f'Erro ao enviar email de notificação: {e}\n')
        else:
            log_error(f'A notificação: "{verb}" não possui recipientes, e por isso não foi enviada.')


auditlog.register(Batch)
auditlog.register(Report)
auditlog.register(ReportFile)
auditlog.register(Invoice)
auditlog.register(DSPProductCodeTranslation)
auditlog.register(Advance)
auditlog.register(AdvanceInvoiceRelation)
